import openpyxl
import json

sheet = openpyxl.load_workbook('./products.xlsx')
products = {}
for sheet_name in sheet.sheetnames:
    sheet = sheet[sheet_name].iter_rows(min_row=2, values_only=True)
    index = 0
    for row in sheet:
        data = {
            "id": row[0],
            "product": row[1],
            "price": row[2],
            "stock": row[3]
        }
        products[index] = data
        index += 1

with open('products.json', 'w') as f:
    f.write(json.dumps(products))

